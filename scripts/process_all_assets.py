#!/usr/bin/env python3
"""
Process Excel, Suno 133 MP3s, and GPT2 29 Images for GeoMelody
"""

import os
import shutil
import openpyxl
import json
import re

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = '/Users/rylinx/Desktop/GeoMelody_Scenic_Spots_Music_Prompts编号.xlsx'
SUNO_SRC_DIR = '/Users/rylinx/Desktop/suno'
GPT2_SRC_DIR = '/Users/rylinx/Desktop/gpt2'

AUDIO_DEST_DIR = os.path.join(ROOT_DIR, 'public', 'audio')
COVERS_DEST_DIR = os.path.join(ROOT_DIR, 'public', 'covers')
DATA_DIR = os.path.join(ROOT_DIR, 'src', 'data')

os.makedirs(AUDIO_DEST_DIR, exist_ok=True)
os.makedirs(COVERS_DEST_DIR, exist_ok=True)

# 1. Copy GPT2 Images
gpt2_files = sorted([f for f in os.listdir(GPT2_SRC_DIR) if not f.startswith('.')])
print(f"Copying {len(gpt2_files)} GPT2 images to {COVERS_DEST_DIR}...")
gpt2_spot_map = {}

for f in gpt2_files:
    src = os.path.join(GPT2_SRC_DIR, f)
    dst = os.path.join(COVERS_DEST_DIR, f)
    shutil.copy2(src, dst)
    
    m = re.match(r"^(\d+)\.(\d+)\.png$", f)
    if m:
        num = int(m.group(1))
        gpt2_spot_map.setdefault(num, []).append(f"/covers/{f}")

print(f"GPT2 Images mapped for spots: {list(gpt2_spot_map.keys())}")

# 2. Process Excel and Copy Suno MP3s
wb = openpyxl.load_workbook(EXCEL_PATH)
sheet = wb.active

header_row_idx = 4
spots_data = []
suno_manifest = {}
suno_src_files = os.listdir(SUNO_SRC_DIR)

copied_audio = 0

for r in range(header_row_idx + 1, sheet.max_row + 1):
    vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    if vals[0] is not None:
        num = int(vals[0])
        spot_id = str(vals[1]).strip()
        name = str(vals[2]).strip()
        en_name = str(vals[3]).strip() if vals[3] else ""
        country = str(vals[4]).strip() if vals[4] else ""
        location = str(vals[5]).strip() if vals[5] else ""
        category = str(vals[6]).strip() if vals[6] else ""
        desc = str(vals[7]).strip() if vals[7] else ""
        tags = str(vals[8]).strip() if vals[8] else ""
        bpm = vals[9] or 72
        scale = str(vals[10]).strip() if vals[10] else ""
        instruments = str(vals[11]).strip() if vals[11] else ""
        natural_sound = str(vals[12]).strip() if vals[12] else ""
        ai_prompt_zh = str(vals[13]).strip() if vals[13] else ""
        ai_prompt_en = str(vals[14]).strip() if vals[14] else ""
        visual_prompt = str(vals[15]).strip() if vals[15] else ""
        filename = str(vals[16]).strip() if len(vals) > 16 and vals[16] else ""

        # Find matching MP3 file
        candidates = [
            f"{filename}.mp3" if filename and not filename.endswith('.mp3') else filename,
            f"{num:02d}.mp3",
            f"{num}.mp3",
            f"{num:03d}.mp3"
        ]
        matched_file = None
        for cand in candidates:
            if cand and cand in suno_src_files:
                matched_file = cand
                break
        if not matched_file:
            for f in suno_src_files:
                if re.match(rf"^0*{num}\.mp3$", f):
                    matched_file = f
                    break

        clean_id = spot_id.replace('-', '_')
        std_audio_name = f"suno_{num:03d}_{clean_id}.mp3"

        if matched_file:
            src_audio = os.path.join(SUNO_SRC_DIR, matched_file)
            dst_std = os.path.join(AUDIO_DEST_DIR, std_audio_name)
            shutil.copy2(src_audio, dst_std)
            # also save as {num}.mp3 and {num:02d}.mp3
            shutil.copy2(src_audio, os.path.join(AUDIO_DEST_DIR, f"{num}.mp3"))
            shutil.copy2(src_audio, os.path.join(AUDIO_DEST_DIR, f"{num:02d}.mp3"))
            copied_audio += 1
        else:
            print(f"Warning: Audio file not found for #{num} {spot_id}")

        tag_list = [t.strip() for t in re.split(r'[,、，|/]', tags) if t.strip()]

        photos = gpt2_spot_map.get(num, [])
        
        spot_obj = {
            "num": num,
            "id": spot_id,
            "name": name,
            "enName": en_name,
            "country": country,
            "location": location,
            "category": category,
            "desc": desc,
            "tags": tag_list,
            "bpm": bpm,
            "scale": scale,
            "instruments": instruments,
            "naturalSound": natural_sound,
            "aiPromptZh": ai_prompt_zh,
            "aiPromptEn": ai_prompt_en,
            "visualPrompt": visual_prompt,
            "audioUrl": f"/audio/{std_audio_name}",
            "photos": photos
        }
        spots_data.append(spot_obj)

        suno_manifest[spot_id] = {
            "num": num,
            "spotId": spot_id,
            "name": name,
            "enName": en_name,
            "country": country,
            "location": location,
            "category": category,
            "url": f"/audio/{std_audio_name}",
            "title": f"{name} · 原生风光律动",
            "enTitle": f"{en_name or name} Soundscape",
            "creator": "Suno AI · GeoMelody",
            "bpm": bpm,
            "scale": scale,
            "instruments": instruments,
            "naturalSound": natural_sound,
            "aiPromptZh": ai_prompt_zh,
            "aiPromptEn": ai_prompt_en
        }

print(f"Copied {copied_audio}/133 audio tracks.")

# Save manifest
manifest_path = os.path.join(DATA_DIR, 'sunoTracksManifest.json')
with open(manifest_path, 'w', encoding='utf-8') as f:
    json.dump(suno_manifest, f, ensure_ascii=False, indent=2)
print(f"Saved manifest to {manifest_path}")

# Output spot data summary
print(f"Total processed spots: {len(spots_data)}")
