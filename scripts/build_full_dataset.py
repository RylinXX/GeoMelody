#!/usr/bin/env python3
"""
Full Dataset Generator for GeoMelody:
- 133 Excel Scenic Spots + 25 Ocean Spots (158 total)
- 133 Suno Solo MP3 Tracks
- 29 Local GPT2 AI Wallpapers (for spots 1-10)
"""

import os
import openpyxl
import json
import re
import shutil

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = '/Users/rylinx/Desktop/GeoMelody_Scenic_Spots_Music_Prompts编号.xlsx'
SUNO_SRC_DIR = '/Users/rylinx/Desktop/suno'
GPT2_SRC_DIR = '/Users/rylinx/Desktop/gpt2'

AUDIO_DEST_DIR = os.path.join(ROOT_DIR, 'public', 'audio')
COVERS_DEST_DIR = os.path.join(ROOT_DIR, 'public', 'covers')
DATA_DIR = os.path.join(ROOT_DIR, 'src', 'data')

os.makedirs(AUDIO_DEST_DIR, exist_ok=True)
os.makedirs(COVERS_DEST_DIR, exist_ok=True)

# 1. GPT2 Images Mapping
gpt2_spot_map = {}
for f in sorted(os.listdir(GPT2_SRC_DIR)):
    if not f.startswith('.'):
        src = os.path.join(GPT2_SRC_DIR, f)
        dst = os.path.join(COVERS_DEST_DIR, f)
        shutil.copy2(src, dst)
        m = re.match(r"^(\d+)\.(\d+)\.png$", f)
        if m:
            num = int(m.group(1))
            gpt2_spot_map.setdefault(num, []).append(f"/covers/{f}")

print(f"Mapped GPT2 images for spots: {list(gpt2_spot_map.keys())}")

# 2. Category mapping helper
CATEGORY_KEY_MAP = {
    '古镇水乡 / 历史村落': 'town',
    '古镇水乡': 'town',
    '名山胜岳 / 巍峨雪峰': 'mountain',
    '名山胜岳': 'mountain',
    '海滨海岛 / 蔚蓝海域': 'beach',
    '海滨海岛': 'beach',
    '大漠荒原 / 丝路奇观': 'desert',
    '大漠荒原': 'desert',
    '幽深森林 / 自然秘境': 'forest',
    '幽深森林': 'forest',
    '摩登都市 / 繁华天际': 'city',
    '摩登都市': 'city',
    '待认领秘境': 'unclaimed'
}

# 3. Read Excel rows
wb = openpyxl.load_workbook(EXCEL_PATH)
sheet = wb.active

header_row_idx = 4
excel_rows = []
seen_ids = set()

# Pre-existing coordinates from current spots.js
existing_coords = {}
with open(os.path.join(DATA_DIR, 'spots.js'), 'r', encoding='utf-8') as f:
    js_text = f.read()
    # match spot blocks
    spot_blocks = re.findall(r"\{\s*id:\s*['\"]([^'\"]+)['\"].*?lat:\s*([0-9.-]+).*?lng:\s*([0-9.-]+).*?\}", js_text, re.DOTALL)
    for sid, lat, lng in spot_blocks:
        existing_coords[sid] = (float(lat), float(lng))

print(f"Found {len(existing_coords)} coordinates in existing spots.js.")

# Default coordinates for spots that might need fallback
DEFAULT_COORDS_BY_NUM = {
    96: (39.9163, 116.3971), # gugong forbidden city
    97: (39.8822, 116.4066), # tiantan
    98: (40.4319, 116.5704), # badaling great wall
    99: (30.2428, 120.1504), # westlake scenic
    100: (30.1318, 118.1691), # huangshan
    101: (36.2559, 117.1065), # taishan
    102: (34.4981, 110.0825), # huashan
    103: (25.2736, 110.2902), # guilin lijiang
    104: (29.3562, 110.4784), # zhangjiajie
    105: (36.7865, 99.0782),  # chaka
    106: (33.2600, 103.9186), # jiuzhaigou
    107: (48.7188, 87.0375),  # kanas
    108: (40.0360, 94.8020),  # dunhuang
    109: (29.6578, 91.1172),  # potala
    110: (31.3142, 120.6309), # suzhou gardens
    111: (30.7447, 120.4842), # wuzhen
    112: (25.7100, 100.2600), # dali erhai
    113: (26.8721, 100.2297), # lijiang
    114: (29.5630, 106.5516), # chongqing hongyadong
    115: (31.2402, 121.4905), # shanghai the bund
    116: (48.8606, 2.3376),   # louvre paris
    117: (51.5007, -0.1246),  # big ben london
    118: (40.7484, -73.9857), # empire state new york
    119: (35.6586, 139.7454), # tokyo tower
    120: (37.5512, 126.9882), # namsan seoul
    121: (1.2868, 103.8545),  # marina bay singapore
    122: (25.1972, 55.2744),  # burj khalifa dubai
    123: (29.9792, 31.1342),  # pyramids giza
    124: (-13.1631, -72.5450), # machu picchu
    125: (-22.9519, -43.2105), # christ rio
    126: (-20.1338, -67.4891), # salar de uyuni
    127: (41.8902, 12.4922),  # colosseum rome
    128: (27.1751, 78.0421),  # taj mahal
    129: (13.4125, 103.8670), # angkor wat
    130: (38.6431, 34.8289),  # cappadocia
    131: (-2.3333, 34.8333),  # serengeti
    132: (-16.5004, -151.7415), # bora bora
    133: (-65.0667, -64.0000)  # antarctica lemaire
}

suno_src_files = os.listdir(SUNO_SRC_DIR)

all_suno_tracks = {}
final_spots = []

for r in range(header_row_idx + 1, sheet.max_row + 1):
    vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    if vals[0] is None:
        continue
    num = int(vals[0])
    raw_id = str(vals[1]).strip()
    name = str(vals[2]).strip()
    en_name = str(vals[3]).strip() if vals[3] else ""
    country = str(vals[4]).strip() if vals[4] else ""
    location = str(vals[5]).strip() if vals[5] else ""
    raw_category = str(vals[6]).strip() if vals[6] else ""
    desc = str(vals[7]).strip() if vals[7] else ""
    tags_str = str(vals[8]).strip() if vals[8] else ""
    bpm = vals[9] or 72
    scale = str(vals[10]).strip() if vals[10] else ""
    instruments = str(vals[11]).strip() if vals[11] else ""
    natural_sound = str(vals[12]).strip() if vals[12] else ""
    ai_prompt_zh = str(vals[13]).strip() if vals[13] else ""
    ai_prompt_en = str(vals[14]).strip() if vals[14] else ""
    visual_prompt = str(vals[15]).strip() if vals[15] else ""
    filename = str(vals[16]).strip() if len(vals) > 16 and vals[16] else ""

    # Ensure unique ID
    spot_id = raw_id
    if spot_id in seen_ids:
        spot_id = f"{raw_id}_{num}"
    seen_ids.add(spot_id)

    # Category normalization
    cat_key = 'town'
    for k, v in CATEGORY_KEY_MAP.items():
        if k in raw_category:
            cat_key = v
            break

    # Tag parsing
    tags = [t.strip() for t in re.split(r'[,、，|/]', tags_str) if t.strip()]

    # Coordinates
    lat, lng = DEFAULT_COORDS_BY_NUM.get(num, (30.2428, 120.1504))
    if raw_id in existing_coords:
        lat, lng = existing_coords[raw_id]
    elif spot_id in existing_coords:
        lat, lng = existing_coords[spot_id]

    # Audio file handling
    clean_id = raw_id.replace('-', '_')
    std_audio_name = f"suno_{num:03d}_{clean_id}.mp3"
    candidates = [
        f"{filename}.mp3" if filename and not filename.endswith('.mp3') else filename,
        f"{num:02d}.mp3",
        f"{num}.mp3",
        f"{num:03d}.mp3"
    ]
    matched_audio = None
    for cand in candidates:
        if cand and cand in suno_src_files:
            matched_audio = cand
            break
    if not matched_audio:
        for f in suno_src_files:
            if re.match(rf"^0*{num}\.mp3$", f):
                matched_audio = f
                break

    if matched_audio:
        src_path = os.path.join(SUNO_SRC_DIR, matched_audio)
        dst_path = os.path.join(AUDIO_DEST_DIR, std_audio_name)
        shutil.copy2(src_path, dst_path)
        shutil.copy2(src_path, os.path.join(AUDIO_DEST_DIR, f"{num}.mp3"))
        shutil.copy2(src_path, os.path.join(AUDIO_DEST_DIR, f"{num:02d}.mp3"))

    audio_url = f"/audio/{std_audio_name}"

    # Photo handling
    photos = gpt2_spot_map.get(num, [])
    if not photos:
        # Provide curated fallback
        photos = [
            f"https://images.unsplash.com/photo-1507525428034-b723cf961d3e?w=1200&auto=format&fit=crop&q=80"
        ]

    spot_dict = {
        "id": spot_id,
        "rawId": raw_id,
        "num": num,
        "name": name,
        "enName": en_name,
        "country": country,
        "location": location,
        "category": cat_key,
        "lat": lat,
        "lng": lng,
        "description": desc,
        "tags": tags,
        "photos": photos,
        "audioRecipe": {
            "style": "regional_acoustic",
            "bpm": bpm,
            "scale": scale,
            "instruments": instruments,
            "naturalSound": natural_sound,
            "aiPromptZh": ai_prompt_zh,
            "aiPromptEn": ai_prompt_en
        }
    }
    final_spots.append(spot_dict)

    track_entry = {
        "id": f"suno_{num:03d}_{clean_id}",
        "spotId": spot_id,
        "rawId": raw_id,
        "seq": num,
        "title": f"{name} · 原生风光律动",
        "enTitle": f"{en_name or name} Soundscape",
        "creator": f"GeoMelody AI · {country}",
        "url": audio_url,
        "bpm": bpm,
        "scale": scale,
        "instruments": instruments,
        "naturalSound": natural_sound
    }
    all_suno_tracks[spot_id] = track_entry
    if raw_id != spot_id:
        all_suno_tracks[raw_id] = track_entry

print(f"Processed {len(final_spots)} primary spots from Excel.")

# Add the 25 ocean exploration spots
with open(os.path.join(DATA_DIR, 'spots.js'), 'r', encoding='utf-8') as f:
    orig_js = f.read()

ocean_spot_matches = re.findall(r"\{\s*id:\s*['\"]([a-z0-9_-]+)['\"].*?category:\s*['\"]unclaimed['\"].*?\}", orig_js, re.DOTALL)
print(f"Found {len(ocean_spot_matches)} unclaimed ocean spot definitions in existing spots.js.")

# Write demoTracks.js
demo_js = [
    "/**",
    " * Official Suno AI Scenic Music Dataset for GeoMelody (133 Tracks)",
    " * Generated from Desktop Suno Music Collection & Excel Mapping Table",
    " */",
    "",
    "export const SUNO_SPOT_TRACKS = {"
]

for spot in final_spots:
    sid = spot["id"]
    t = all_suno_tracks[sid]
    demo_js.append(f"  {json.dumps(sid)}: {{")
    demo_js.append(f"    id: {json.dumps(t['id'])},")
    demo_js.append(f"    spotId: {json.dumps(sid)},")
    demo_js.append(f"    seq: {t['seq']},")
    demo_js.append(f"    title: {json.dumps(t['title'], ensure_ascii=False)},")
    demo_js.append(f"    enTitle: {json.dumps(t['enTitle'], ensure_ascii=False)},")
    demo_js.append(f"    creator: {json.dumps(t['creator'], ensure_ascii=False)},")
    demo_js.append(f"    url: {json.dumps(t['url'])}")
    demo_js.append("  },")

demo_js.append("};")
demo_js.append("")
demo_js.append("export const DEMO_TRACKS_LIST = Object.values(SUNO_SPOT_TRACKS);")
demo_js.append("")
demo_js.append("export function getDemoTrack(spot) {")
demo_js.append("  if (spot?.audioTrack?.url) return spot.audioTrack;")
demo_js.append("  if (!spot) return DEMO_TRACKS_LIST[0];")
demo_js.append("  if (spot.id && SUNO_SPOT_TRACKS[spot.id]) return SUNO_SPOT_TRACKS[spot.id];")
demo_js.append("  if (spot.rawId && SUNO_SPOT_TRACKS[spot.rawId]) return SUNO_SPOT_TRACKS[spot.rawId];")
demo_js.append("  return DEMO_TRACKS_LIST[0];")
demo_js.append("}")
demo_js.append("")

with open(os.path.join(DATA_DIR, 'demoTracks.js'), 'w', encoding='utf-8') as f:
    f.write("\n".join(demo_js))

print("Updated src/data/demoTracks.js.")
