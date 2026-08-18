#!/usr/bin/env python3
"""
Generate clean, valid src/data/spots.js
"""

import os
import openpyxl
import json
import re

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = '/Users/rylinx/Desktop/GeoMelody_Scenic_Spots_Music_Prompts编号.xlsx'
GPT2_SRC_DIR = '/Users/rylinx/Desktop/gpt2'
DATA_DIR = os.path.join(ROOT_DIR, 'src', 'data')

# 1. GPT2 Images Mapping
gpt2_spot_map = {}
for f in sorted(os.listdir(GPT2_SRC_DIR)):
    if not f.startswith('.'):
        m = re.match(r"^(\d+)\.(\d+)\.png$", f)
        if m:
            num = int(m.group(1))
            gpt2_spot_map.setdefault(num, []).append(f"/covers/{f}")

CATEGORY_KEY_MAP = {
    '古镇水乡 / 历史村落': 'town',
    '古镇水乡': 'town',
    '名山胜岳 / 巍峨雪峰': 'mountain',
    '名山胜岳': 'mountain',
    '海滨海岛 / 蔚蓝海域': 'beach',
    '海滨海岛': 'beach',
    '大漠荒原 / 丝路奇观': 'desert',
    '大漠荒原': 'desert',
    '幽深森林 / 自然秘境': 'forest',
    '幽深森林': 'forest',
    '摩登都市 / 繁华天际': 'city',
    '摩登都市': 'city',
    '待认领秘境': 'unclaimed'
}

# 25 Unclaimed Ocean Spot Definitions
OCEAN_SPOTS = [
  {
    "id": "mariana-trench",
    "name": "马里亚纳海沟 · 挑战者深渊",
    "enName": "Mariana Trench · Challenger Deep",
    "location": "西太平洋 · 关岛西南海域",
    "country": "西太平洋公海",
    "category": "unclaimed",
    "isUnclaimed": True,
    "lat": 11.3493,
    "lng": 142.1996,
    "description": "【待认领秘境】地球最深处（-10,994米）。幽深莫测的冥界洋底，水压超千个大气压，等待第一位勇者命名认领与谱写专属音律。",
    "tags": ["待认领", "地球极深", "挑战者深渊", "深海冥界", "神秘秘境"],
    "photos": [
      "https://images.unsplash.com/photo-1544551763-46a013bb70d5?auto=format&fit=crop&w=1920&q=85",
      "https://images.unsplash.com/photo-1518837695005-2083093ee35b?auto=format&fit=crop&w=1920&q=85"
    ],
    "audioRecipe": {
      "style": "deep_ocean_ambient",
      "bpm": 52,
      "scale": "Deep Abyss Drone · 432Hz",
      "instruments": "深海水听器共鸣 · 极低频波纹 · 鲸歌远鸣",
      "naturalSound": "waves",
      "prompt": "深海万米深渊环境音、水下低频共振、神秘鲸鸣、深邃治愈助眠"
    }
  },
  {
    "id": "point-nemo",
    "name": "尼莫点 · 太平洋海洋难抵极",
    "enName": "Point Nemo · Oceanic Pole of Inaccessibility",
    "location": "南太平洋 · 航天器重返墓地",
    "country": "南太平洋公海",
    "category": "unclaimed",
    "isUnclaimed": True,
    "lat": -48.8767,
    "lng": -123.3933,
    "description": "【待认领秘境】地球上距离任何陆地最遥远的点（2,688公里）。最近的人类往往是头顶400公里呼啸而过的国际空间站宇航员。万籁俱寂的浩瀚大洋，等待探险者留下足迹。",
    "tags": ["待认领", "海洋难抵极", "航天器坟场", "绝对孤独", "深洋绝境"],
    "photos": [
      "https://images.unsplash.com/photo-1505118380757-91f5f5632de0?auto=format&fit=crop&w=1920&q=85",
      "https://images.unsplash.com/photo-1498084393753-b411b2d26b34?auto=format&fit=crop&w=1920&q=85"
    ],
    "audioRecipe": {
      "style": "deep_ocean_ambient",
      "bpm": 48,
      "scale": "Infinite Solitude · 528Hz",
      "instruments": "微风泛音合成器 · 孤舟水波 · 远古潮汐",
      "naturalSound": "waves",
      "prompt": "无垠太平洋、深邃孤寂、远方微弱海浪声、静心冥想与深度睡眠"
    }
  },
  {
    "id": "bermuda-triangle",
    "name": "百慕大神秘三角 · 罗盘静默海域",
    "enName": "Bermuda Triangle · Sargasso Western Apex",
    "location": "北大西洋 · 萨尔加斯海西界",
    "country": "大西洋公海",
    "category": "unclaimed",
    "isUnclaimed": True,
    "lat": 25.0000,
    "lng": -71.0000,
    "description": "【待认领秘境】大西洋上传奇的神秘海域，幽蓝无垠的洋流与传说中的时空迷雾，等待属于你的探险音符与传奇故事。",
    "tags": ["待认领", "百慕大", "未解之谜", "幽蓝洋流", "大西洋秘境"],
    "photos": [
      "https://images.unsplash.com/photo-1518837695005-2083093ee35b?auto=format&fit=crop&w=1920&q=85",
      "https://images.unsplash.com/photo-1507525428034-b723cf961d3e?auto=format&fit=crop&w=1920&q=85"
    ],
    "audioRecipe": {
      "style": "deep_ocean_ambient",
      "bpm": 60,
      "scale": "Mysterious Horizon · E Minor",
      "instruments": "神秘长音合成器 · 罗盘微响 · 潮涌脉冲",
      "naturalSound": "wind",
      "prompt": "神秘百慕大海域、微弱电磁脉冲泛音、海浪与海风、沉浸式冥想氛围"
    }
  },
  {
    "id": "easter-island-abyss",
    "name": "复活节岛外海深海断崖",
    "enName": "Easter Island Offshore Abyss",
    "location": "东南太平洋 · 拉帕努伊海沟",
    "country": "智利外海",
    "category": "unclaimed",
    "isUnclaimed": True,
    "lat": -27.1127,
    "lng": -109.3497,
    "description": "【待认领秘境】遥远太平洋孤岛外延数千米的深海断层，摩艾巨石注视的远海，蕴藏着古老玻利尼西亚航海史诗。",
    "tags": ["待认领", "复活节岛", "太平洋孤岛", "深海断崖", "远洋遗迹"],
    "photos": ["https://images.unsplash.com/photo-1507525428034-b723cf961d3e?auto=format&fit=crop&w=1920&q=85"],
    "audioRecipe": { "style": "deep_ocean_ambient", "bpm": 56, "scale": "Ancient Pacific Chant", "instruments": "海螺鸣响 · 沉吟低音 · 浪涌", "naturalSound": "waves", "prompt": "古老太平洋波涛、沉吟悠远海螺音、深海冥想" }
  },
  {
    "id": "galapagos-rift",
    "name": "加拉帕戈斯深海海底热泉",
    "enName": "Galapagos Rift Hydrothermal Vents",
    "location": "东赤道太平洋 · 洋中脊断裂带",
    "country": "厄瓜多尔外海",
    "category": "unclaimed",
    "isUnclaimed": True,
    "lat": 0.7983,
    "lng": -86.1500,
    "description": "【待认领秘境】人类历史上首次发现黑烟囱海底热泉生命绿洲的科学圣地，黑暗深洋中沸腾的生命奇迹。",
    "tags": ["待认领", "海底热泉", "生命起源", "黑烟囱", "深海探奇"],
    "photos": ["https://images.unsplash.com/photo-1544551763-46a013bb70d5?auto=format&fit=crop&w=1920&q=85"],
    "audioRecipe": { "style": "deep_ocean_ambient", "bpm": 64, "scale": "Primordial Heat", "instruments": "热液气泡脉动 · 舒缓铜锣 · 晶莹泛音", "naturalSound": "water", "prompt": "海底热泉生命脉动、温暖深海水声、空灵治愈" }
  },
  {
    "id": "tahiti-blue-hole",
    "name": "大溪地群岛外海无底蓝洞",
    "enName": "Tahiti Outer Abyss Blue Hole",
    "location": "南太平洋 · 土阿莫土珊瑚环礁外缘",
    "country": "法属波利尼西亚外海",
    "category": "unclaimed",
    "isUnclaimed": True,
    "lat": -15.0000,
    "lng": -146.0000,
    "description": "【待认领秘境】从浅蓝绿松石珊瑚泻湖骤降至数千米幽暗深蓝的垂直绝壁，潜水员与探险家的终极朝圣之境。",
    "tags": ["待认领", "大溪地深洋", "海洋蓝洞", "深海悬崖", "纯净琉璃"],
    "photos": ["https://images.unsplash.com/photo-1507525428034-b723cf961d3e?auto=format&fit=crop&w=1920&q=85"],
    "audioRecipe": { "style": "deep_ocean_ambient", "bpm": 58, "scale": "Azure Descent", "instruments": "水晶琴 · 柔和水滴 · 悠长和弦", "naturalSound": "water", "prompt": "蔚蓝深海下沉、水晶空灵回响、纯净宁静" }
  },
  {
    "id": "mauritius-underwater-waterfall",
    "name": "毛里求斯海底瀑布幻境",
    "enName": "Mauritius Underwater Waterfall Trench",
    "location": "西南印度洋 · 勒莫恩海角外缘",
    "country": "毛里求斯外海",
    "category": "unclaimed",
    "isUnclaimed": True,
    "lat": -20.4683,
    "lng": 57.3094,
    "description": "【待认领秘境】印度洋海床泥沙沿海沟倾泻而下形成的壮观视觉奇迹，大自然的鬼斧神工宛若水下飞瀑直入地心。",
    "tags": ["待认领", "海底瀑布", "印度洋奇观", "勒莫恩角", "幻境洋流"],
    "photos": ["https://images.unsplash.com/photo-1507525428034-b723cf961d3e?auto=format&fit=crop&w=1920&q=85"],
    "audioRecipe": { "style": "deep_ocean_ambient", "bpm": 68, "scale": "Oceanic Cascade", "instruments": "流水层叠乐 · 竖琴泛音 · 暖色合成器", "naturalSound": "water", "prompt": "海底奔腾水流声、温柔舒缓水滴竖琴、清新灵动" }
  },
  {
    "id": "java-trench",
    "name": "爪哇海沟 · 巽他极深渊",
    "enName": "Java Trench · Sunda Deep",
    "location": "东印度洋 · 印度尼西亚南外海",
    "country": "印度洋公海",
    "category": "unclaimed",
    "isUnclaimed": True,
    "lat": -10.3167,
    "lng": 109.9667,
    "description": "【待认领秘境】印度洋最深处（-7,290米），环太平洋火山地震带的深海前哨，见证板块撞击的洪荒原力。",
    "tags": ["待认领", "爪哇海沟", "印度洋极深", "板块裂隙", "幽暗深渊"],
    "photos": ["https://images.unsplash.com/photo-1518837695005-2083093ee35b?auto=format&fit=crop&w=1920&q=85"],
    "audioRecipe": { "style": "deep_ocean_ambient", "bpm": 50, "scale": "Subduction Drone", "instruments": "极低音大提琴 · 远方海鸣 · 舒缓白噪音", "naturalSound": "waves", "prompt": "深邃地幔共振音、远洋波浪、静心助眠" }
  },
  {
    "id": "drake-passage",
    "name": "德雷克海峡狂暴西风带",
    "enName": "Drake Passage Furious Fifties",
    "location": "南大洋 · 合恩角与南极半岛之间",
    "country": "南大洋公海",
    "category": "unclaimed",
    "isUnclaimed": True,
    "lat": -58.5000,
    "lng": -65.0000,
    "description": "【待认领秘境】全世界最危险也是最壮观的航道，咆哮西风带与极地狂涌在此交汇，通向南极大陆的勇敢者阶梯。",
    "tags": ["待认领", "德雷克海峡", "咆哮西风带", "南大洋狂浪", "极地航线"],
    "photos": ["https://images.unsplash.com/photo-1498084393753-b411b2d26b34?auto=format&fit=crop&w=1920&q=85"],
    "audioRecipe": { "style": "deep_ocean_ambient", "bpm": 62, "scale": "Antarctic Storm", "instruments": "极地风声合成器 · 冰浪拍击 · 史诗交响泛音", "naturalSound": "wind", "prompt": "南极狂风与海浪、辽阔壮美、史诗纯乐环境音" }
  },
  {
    "id": "azores-rift",
    "name": "亚速尔群岛大西洋中脊三叉口",
    "enName": "Azores Triple Junction Mid-Atlantic Ridge",
    "location": "中大西洋 · 欧亚/北美/非洲板块三联点",
    "country": "葡萄牙外海",
    "category": "unclaimed",
    "isUnclaimed": True,
    "lat": 38.5000,
    "lng": -28.0000,
    "description": "【待认领秘境】三大板块在海面下缓缓撕裂分离的壮丽洋底山脊，传说中亚特兰蒂斯沉睡之地。",
    "tags": ["待认领", "亚速尔", "大西洋洋中脊", "板块三联点", "大西洲传说"],
    "photos": ["https://images.unsplash.com/photo-1507525428034-b723cf961d3e?auto=format&fit=crop&w=1920&q=85"],
    "audioRecipe": { "style": "deep_ocean_ambient", "bpm": 55, "scale": "Atlantis Melody", "instruments": "古典鲁特琴 · 远方唱咏 · 温暖洋流", "naturalSound": "waves", "prompt": "大西洋中脊神秘水声、温暖洋流和弦、宁静空灵" }
  }
]

# Coordinates mapping
DEFAULT_COORDS_BY_NUM = {
    96: (39.9163, 116.3971), 97: (39.8822, 116.4066), 98: (40.4319, 116.5704), 99: (30.2428, 120.1504),
    100: (30.1318, 118.1691), 101: (36.2559, 117.1065), 102: (34.4981, 110.0825), 103: (25.2736, 110.2902),
    104: (29.3562, 110.4784), 105: (36.7865, 99.0782), 106: (33.2600, 103.9186), 107: (48.7188, 87.0375),
    108: (40.0360, 94.8020), 109: (29.6578, 91.1172), 110: (31.3142, 120.6309), 111: (30.7447, 120.4842),
    112: (25.7100, 100.2600), 113: (26.8721, 100.2297), 114: (29.5630, 106.5516), 115: (31.2402, 121.4905),
    116: (48.8606, 2.3376), 117: (51.5007, -0.1246), 118: (40.7484, -73.9857), 119: (35.6586, 139.7454),
    120: (37.5512, 126.9882), 121: (1.2868, 103.8545), 122: (25.1972, 55.2744), 123: (29.9792, 31.1342),
    124: (-13.1631, -72.5450), 125: (-22.9519, -43.2105), 126: (-20.1338, -67.4891), 127: (41.8902, 12.4922),
    128: (27.1751, 78.0421), 129: (13.4125, 103.8670), 130: (38.6431, 34.8289), 131: (-2.3333, 34.8333),
    132: (-16.5004, -151.7415), 133: (-65.0667, -64.0000)
}

# Known coordinates for spots 1-95
KNOWN_COORDS = {
    'wuzhen': (30.7447, 120.4842), 'zhouzhuang': (31.1158, 120.8447), 'hongcun': (30.0033, 117.9892),
    'fenghuang': (27.9542, 109.6006), 'lijiang': (26.8721, 100.2297), 'nanxun': (30.8753, 120.4206),
    'xitang': (30.9442, 120.8903), 'wuyuan': (29.2547, 117.8611), 'pingyao': (37.2022, 112.1794),
    'kyoto': (35.0116, 135.6778), 'shirakawago': (36.2575, 136.9064), 'hallstatt': (47.5622, 13.6493),
    'florence': (43.7696, 11.2558), 'meili': (28.4500, 98.7833), 'yulong': (27.0983, 100.2033),
    'kailash': (31.0667, 81.3125), 'yading': (28.3833, 100.3333), 'gongga': (29.5833, 101.8833),
    'siguniang': (31.1000, 102.9000), 'everest': (27.9881, 86.9250), 'fuji': (35.3606, 138.7274),
    'matterhorn': (45.9763, 7.6586), 'lofoten': (68.1667, 13.7500), 'sanya': (18.2528, 109.5119),
    'weizhou': (21.0458, 109.1172), 'maldives': (3.2028, 73.2207), 'bali': (-8.4095, 115.1889),
    'santorini': (36.3932, 25.4615), 'semporna': (4.4817, 118.6111), 'borabora': (-16.5004, -151.7415),
    'dunhuang': (40.0360, 94.8020), 'shapotou': (37.5028, 105.0278), 'taklamakan': (38.9000, 83.6500),
    'kumtag': (41.2833, 93.4333), 'sahara': (23.4162, 25.6628), 'wadirum': (29.5736, 35.4208),
    'zhangjiajie': (29.3562, 110.4784), 'jiuzhaigou': (33.2600, 103.9186), 'changbai': (42.0000, 128.0500),
    'yakushima': (30.3585, 130.5286), 'blackforest': (48.0000, 8.2000), 'redwood': (41.2132, -124.0046),
    'plitvice': (44.8654, 15.5820), 'shanghai': (31.2402, 121.4905), 'tokyo': (35.6762, 139.6503),
    'newyork': (40.7128, -74.0060), 'westlake': (30.2428, 120.1504), 'chaka': (36.7865, 99.0782),
    'guilin': (25.2736, 110.2902), 'gugong': (39.9163, 116.3971), 'potala': (29.6578, 91.1172),
    'taishan': (36.2559, 117.1065), 'huangshan': (30.1318, 118.1691), 'kanas': (48.7188, 87.0375),
    'tajmahal': (27.1751, 78.0421), 'angkorwat': (13.4125, 103.8670), 'machupicchu': (-13.1631, -72.5450),
    'cappadocia': (38.6431, 34.8289), 'colosseum': (41.8902, 12.4922), 'greatbarrierreef': (-18.2871, 147.6992),
    'sydneyoperahouse': (-33.8568, 151.2153), 'louvremuseum': (48.8606, 2.3376), 'salardeuyuni': (-20.1338, -67.4891)
}

wb = openpyxl.load_workbook(EXCEL_PATH)
sheet = wb.active

header_row_idx = 4
spots_data = []
seen_ids = set()

for r in range(header_row_idx + 1, sheet.max_row + 1):
    vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    if vals[0] is None:
        continue
    num = int(vals[0])
    raw_id = str(vals[1]).strip()
    name = str(vals[2]).strip()
    en_name = str(vals[3]).strip() if vals[3] else ""
    country = str(vals[4]).strip() if vals[4] else ""
    location = str(vals[5]).strip() if vals[5] else ""
    raw_category = str(vals[6]).strip() if vals[6] else ""
    desc = str(vals[7]).strip() if vals[7] else ""
    tags_str = str(vals[8]).strip() if vals[8] else ""
    bpm = vals[9] or 72
    scale = str(vals[10]).strip() if vals[10] else ""
    instruments = str(vals[11]).strip() if vals[11] else ""
    natural_sound = str(vals[12]).strip() if vals[12] else ""
    ai_prompt_zh = str(vals[13]).strip() if vals[13] else ""
    ai_prompt_en = str(vals[14]).strip() if vals[14] else ""

    spot_id = raw_id
    if spot_id in seen_ids:
        spot_id = f"{raw_id}_{num}"
    seen_ids.add(spot_id)

    cat_key = 'town'
    for k, v in CATEGORY_KEY_MAP.items():
        if k in raw_category:
            cat_key = v
            break

    tags = [t.strip() for t in re.split(r'[,、，|/]', tags_str) if t.strip()]

    # Coords lookup
    lat, lng = (30.2428, 120.1504)
    if num in DEFAULT_COORDS_BY_NUM:
        lat, lng = DEFAULT_COORDS_BY_NUM[num]
    elif raw_id in KNOWN_COORDS:
        lat, lng = KNOWN_COORDS[raw_id]

    # Photos
    photos = gpt2_spot_map.get(num, [])
    if not photos:
        photos = [
            f"https://images.unsplash.com/photo-1507525428034-b723cf961d3e?w=1200&auto=format&fit=crop&q=80"
        ]

    spot_obj = {
        "id": spot_id,
        "name": name,
        "enName": en_name or name,
        "location": location,
        "country": country,
        "category": cat_key,
        "lat": lat,
        "lng": lng,
        "description": desc,
        "tags": tags,
        "photos": photos,
        "audioRecipe": {
            "style": "regional_acoustic",
            "bpm": int(bpm) if isinstance(bpm, (int, float)) else 72,
            "scale": scale,
            "instruments": instruments,
            "naturalSound": natural_sound,
            "aiPromptZh": ai_prompt_zh,
            "aiPromptEn": ai_prompt_en
        }
    }
    if spot_id != raw_id:
        spot_obj["rawId"] = raw_id
    spots_data.append(spot_obj)

# Append ocean spots
for os_spot in OCEAN_SPOTS:
    spots_data.append(os_spot)

print(f"Total compiled spots: {len(spots_data)}")

# Write to spots.js
js_content = "/**\n * GeoMelody Global Scenic Spots Dataset\n * Total: 133 Excel Scenic Spots + 10 Ocean Exploration Spots = 143 Master Spots\n */\n\nexport const SCENIC_SPOTS = " + json.dumps(spots_data, ensure_ascii=False, indent=2) + ";\n"

with open(os.path.join(DATA_DIR, 'spots.js'), 'w', encoding='utf-8') as f:
    f.write(js_content)

print("Generated valid src/data/spots.js successfully.")
