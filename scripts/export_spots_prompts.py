import json
import subprocess
import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Fetch full spots data and matched tracks using Node.js
node_cmd = """node -e "
import('./src/data/spots.js').then(async s => {
  const dt = await import('./src/data/demoTracks.js');
  const result = s.SCENIC_SPOTS.map((spot, idx) => {
    const track = dt.getDemoTrack(spot);
    const isDirect = !!dt.SUNO_SPOT_TRACKS[spot.id];
    const filename = track.url ? track.url.replace('/audio/', '') : '';
    return {
      index: idx + 1,
      id: spot.id,
      name: spot.name,
      enName: spot.enName || '',
      country: spot.country || '',
      location: spot.location || '',
      lng: spot.lng,
      lat: spot.lat,
      category: spot.category || 'town',
      description: spot.description || '',
      tags: (spot.tags || []).join('、'),
      audioRecipe: spot.audioRecipe || {},
      trackTitle: track.title || '',
      trackEnTitle: track.enTitle || '',
      trackCreator: track.creator || '',
      matchType: isDirect ? '专属定制曲目' : '分类意境伴奏',
      audioUrl: track.url || '',
      audioFile: filename
    };
  });
  console.log(JSON.stringify(result));
});
" """

res = subprocess.run(node_cmd, shell=True, capture_output=True, text=True, cwd=os.getcwd())
if res.returncode != 0:
    raise Exception(f"Node execution failed: {res.stderr}")

spots_data = json.loads(res.stdout)
print(f"Successfully loaded {len(spots_data)} scenic spots with matched tracks.")

CATEGORY_MAP = {
    'town': '古镇水乡 / 历史村落',
    'mountain': '名山胜岳 / 巍峨雪峰',
    'island': '海滨海岛 / 蔚蓝海域',
    'desert': '浩瀚大漠 / 丝路遗迹',
    'forest': '幽深森林 / 自然秘境',
    'city': '都会风情 / 人文地标',
    'lake': '宁静湖泊 / 高原圣湖',
    'historic': '历史奇迹 / 古代文明',
    'unclaimed': '待认领秘境 / 探索坐标'
}

sound_map_zh = {
    'rain': '烟雨滴檐 / 细雨淅淅',
    'wind': '山间清风 / 微风拂过',
    'waves': '海浪拍岸 / 潮汐涌动',
    'stream': '潺潺流水 / 溪流清音',
    'birds': '晨曦鸟鸣 / 空山鸟语',
    'bell': '古刹钟声 / 远山清钟',
    'night': '夏夜虫鸣 / 静谧夜色'
}

sound_map_en = {
    'rain': 'gentle rain ambience, raindrops on roof',
    'wind': 'soft breeze, mountain wind ambience',
    'waves': 'gentle ocean waves, rhythmic tides',
    'stream': 'flowing stream, water trickle soundscape',
    'birds': 'morning bird singing, forest birds chirp',
    'bell': 'distant temple bell toll, meditation chime',
    'night': 'crickets at night, calm evening atmosphere'
}

en_style_map = {
    'town': 'traditional oriental acoustic folk, guzheng, bamboo flute, acoustic guitar, peaceful water town vibe, relaxing, serene',
    'mountain': 'cinematic ambient, grand mountain landscape, epic yet calm, airy bamboo flute, handpan, strings, ethereal atmosphere',
    'island': 'tropical chillout, acoustic guitar, smooth piano, ukulele, ocean wave ambient, peaceful seaside sunset, warm breeze',
    'desert': 'silk road world music, oud, duduk, middle eastern hand percussion, desert wind, mysterious, atmospheric journey',
    'forest': 'healing nature ambient, acoustic guitar fingerstyle, bird song, wind chimes, lush green forest, zen meditation',
    'city': 'lo-fi chillhop, jazzy piano chords, mellow synth pads, urban coffee shop vibe, nostalgic, smooth groove',
    'lake': 'crystal clear piano, ambient reverb pads, tranquil lake reflection, gentle acoustic strings, deep relaxation, 432hz',
    'historic': 'ancient heritage soundtrack, ethnic instruments, monumental hall reverb, historical grandeur, poetic and timeless',
    'unclaimed': 'ambient atmospheric drone, deep ocean chillout, ethereal handpan, crystal clear synth pads, peaceful wilderness'
}

# 2. Enrich rows
enriched_rows = []
for item in spots_data:
    category = item['category']
    category_zh = CATEGORY_MAP.get(category, '自然人文')
    recipe = item['audioRecipe']
    bpm = recipe.get('bpm', 72)
    scale = recipe.get('scale', '五声调式 · 432Hz')
    instruments = recipe.get('instruments', '原声吉他 · 氛围合成器 · 空间混响')
    natural_sound = recipe.get('naturalSound', 'wind')
    existing_prompt = recipe.get('prompt', '')

    sound_zh = sound_map_zh.get(natural_sound, '环境声场')
    sound_en = sound_map_en.get(natural_sound, 'natural ambient soundscape')

    zh_prompt_keywords = f"国风氛围、{item['name']}、{category_zh}、{instruments}、BPM {bpm}、{scale}、{sound_zh}、空灵舒缓、冥想疗愈、纯器乐无歌词、电影级空间混响、432Hz共鸣"
    if existing_prompt:
        zh_prompt_keywords = f"{existing_prompt}、BPM {bpm}、{scale}、{sound_zh}、高清无损立体声"

    base_en_style = en_style_map.get(category, 'ambient acoustic chillout')
    en_prompt_keywords = f"{base_en_style}, instrumental, no vocals, {bpm} BPM, {sound_en}, pristine spatial reverb, emotional, cinematic travel vlog background music"
    visual_prompt = f"Cinematic photography of {item['enName']} ({item['name']}), {item['location']}, breathtaking view, {item['tags'].replace('、', ', ')}, golden hour lighting, atmospheric perspective, 8k resolution, photorealistic, masterpiece, depth of field"

    enriched_rows.append({
        'index': item['index'],
        'id': item['id'],
        'name': item['name'],
        'en_name': item['enName'],
        'country': item['country'],
        'location': item['location'],
        'lng': item['lng'],
        'lat': item['lat'],
        'category_zh': category_zh,
        'description': item['description'],
        'tags': item['tags'],
        'bpm': bpm,
        'scale': scale,
        'instruments': instruments,
        'soundscape': sound_zh,
        'zh_prompt': zh_prompt_keywords,
        'en_prompt': en_prompt_keywords,
        'visual_prompt': visual_prompt,
        'track_title': item['trackTitle'],
        'track_creator': item['trackCreator'],
        'match_type': item['matchType'],
        'audio_file': item['audioFile']
    })

# 3. Headers Configuration (Last column is "对应的音乐文件")
headers = [
    ("序号", 6),
    ("点位ID", 16),
    ("点位名称 (中文)", 22),
    ("英文名称 (English)", 26),
    ("所属国家", 12),
    ("地理位置 / 省市", 22),
    ("经度 (Lng)", 12),
    ("纬度 (Lat)", 12),
    ("景观分类", 18),
    ("点位意境与人文描述", 45),
    ("核心特色标签", 25),
    ("推荐BPM", 10),
    ("音乐调式/律动", 20),
    ("核心配器音色", 28),
    ("自然声场采样", 18),
    ("AI音乐生成提示词 (中文关键词)", 50),
    ("AI Music Prompt (English)", 55),
    ("Midjourney视觉提示词 (Visual Prompt)", 50),
    ("对照音乐曲名", 24),
    ("音乐创作者/艺术家", 22),
    ("音乐匹配类型", 16),
    ("对应的音乐文件", 28)
]

# 4. Create Styled Excel Workbook with openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "GeoMelody点位与对照音乐名录"
ws.views.sheetView[0].showGridLines = True

# Color Theme: Premium Slate / Cyan Theme
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # slate-800
header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
sub_header_font = Font(name="Microsoft YaHei", size=9, italic=True, color="94A3B8")

title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # slate-900
title_font = Font(name="Microsoft YaHei", size=15, bold=True, color="38BDF8") # Cyan-400

# Write Title Banner (Row 1-2)
ws.merge_cells("A1:V1")
title_cell = ws["A1"]
title_cell.value = "🎵 GeoMelody 全球地理音乐胜景 · 地图点位与对照音乐文件全量清单"
title_cell.font = title_font
title_cell.fill = title_fill
title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 42

ws.merge_cells("A2:V2")
sub_cell = ws["A2"]
sub_cell.value = f"包含当前地图全部 {len(enriched_rows)} 个胜景点位坐标、人文意境描述、推荐音色调式、AI提示词、对照音乐曲目及对应音乐音频文件 | 导出时间：2026年8月"
sub_cell.font = sub_header_font
sub_cell.fill = title_fill
sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 24

# Write Table Headers (Row 4)
header_row_idx = 4
ws.row_dimensions[header_row_idx].height = 32

for col_idx, (header_text, width) in enumerate(headers, start=1):
    cell = ws.cell(row=header_row_idx, column=col_idx, value=header_text)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = width

# Borders & Styles
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)
zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # slate-50

data_font = Font(name="Microsoft YaHei", size=10, color="0F172A")
id_font = Font(name="Consolas", size=9, color="475569")
audio_file_font = Font(name="Consolas", size=9, bold=True, color="0284C7") # Cyan-600
prompt_font = Font(name="Microsoft YaHei", size=9, color="0369A1")
en_prompt_font = Font(name="Segoe UI", size=9, color="334155")
track_font = Font(name="Microsoft YaHei", size=10, bold=True, color="047857") # Emerald-700

# Populate Data Rows
for idx, item in enumerate(enriched_rows, start=1):
    row_idx = header_row_idx + idx
    ws.row_dimensions[row_idx].height = 36
    
    row_data = [
        item['index'],
        item['id'],
        item['name'],
        item['en_name'],
        item['country'],
        item['location'],
        item['lng'],
        item['lat'],
        item['category_zh'],
        item['description'],
        item['tags'],
        item['bpm'],
        item['scale'],
        item['instruments'],
        item['soundscape'],
        item['zh_prompt'],
        item['en_prompt'],
        item['visual_prompt'],
        item['track_title'],
        item['track_creator'],
        item['match_type'],
        item['audio_file'] # 最后一列：对应的音乐文件
    ]
    
    is_zebra = (idx % 2 == 0)
    for col_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.border = thin_border
        
        # Alignments & Fonts
        if col_idx in (1, 7, 8, 12): # Index, Lng, Lat, BPM
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = data_font
        elif col_idx == 2: # Spot ID
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.font = id_font
        elif col_idx in (3, 5, 9, 15, 21): # Name, Country, Category, Soundscape, Match Type
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = data_font
        elif col_idx in (16,): # Chinese Prompt
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.font = prompt_font
        elif col_idx in (17, 18): # English Prompts
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.font = en_prompt_font
        elif col_idx == 19: # Track Title
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.font = track_font
        elif col_idx == 22: # Audio File (Last Column)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.font = audio_file_font
        else:
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.font = data_font
            
        if is_zebra:
            c.fill = zebra_fill

# Freeze Pane below headers
ws.freeze_panes = "A5"

# Save Excel Output
excel_path = os.path.join(os.getcwd(), "GeoMelody_Scenic_Spots_Music_Prompts.xlsx")
wb.save(excel_path)
print(f"Excel file saved successfully to: {excel_path}")

# Save CSV Output with UTF-8 BOM
csv_path = os.path.join(os.getcwd(), "GeoMelody_Scenic_Spots_Music_Prompts.csv")
with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
    writer = csv.writer(f)
    writer.writerow([h[0] for h in headers])
    for item in enriched_rows:
        writer.writerow([
            item['index'],
            item['id'],
            item['name'],
            item['en_name'],
            item['country'],
            item['location'],
            item['lng'],
            item['lat'],
            item['category_zh'],
            item['description'],
            item['tags'],
            item['bpm'],
            item['scale'],
            item['instruments'],
            item['soundscape'],
            item['zh_prompt'],
            item['en_prompt'],
            item['visual_prompt'],
            item['track_title'],
            item['track_creator'],
            item['match_type'],
            item['audio_file'] # 最后一列：对应的音乐文件
        ])
print(f"CSV file saved successfully to: {csv_path}")
