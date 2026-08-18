#!/usr/bin/env python3
"""
Generate clean, complete src/data/spots.js
"""

import os
import openpyxl
import json
import re

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = '/Users/rylinx/Desktop/GeoMelody_Scenic_Spots_Music_Prompts编号.xlsx'
GPT2_SRC_DIR = '/Users/rylinx/Desktop/gpt2'
DATA_DIR = os.path.join(ROOT_DIR, 'src', 'data')

# 1. GPT2 Images Mapping
gpt2_spot_map = {}
for f in sorted(os.listdir(GPT2_SRC_DIR)):
    if not f.startswith('.'):
        m = re.match(r"^(\d+)\.(\d+)\.png$", f)
        if m:
            num = int(m.group(1))
            gpt2_spot_map.setdefault(num, []).append(f"/covers/{f}")

CATEGORY_KEY_MAP = {
    '古镇水乡 / 历史村落': 'town',
    '古镇水乡': 'town',
    '名山胜岳 / 巍峨雪峰': 'mountain',
    '名山胜岳': 'mountain',
    '海滨海岛 / 蔚蓝海域': 'beach',
    '海滨海岛': 'beach',
    '大漠荒原 / 丝路奇观': 'desert',
    '大漠荒原': 'desert',
    '幽深森林 / 自然秘境': 'forest',
    '幽深森林': 'forest',
    '摩登都市 / 繁华天际': 'city',
    '摩登都市': 'city',
    '待认领秘境': 'unclaimed'
}

# Pre-existing coordinates and photos from current spots.js
existing_spots_map = {}
with open(os.path.join(DATA_DIR, 'spots.js'), 'r', encoding='utf-8') as f:
    js_text = f.read()

# Extract spot objects using regex / JS parsing
# Let's extract existing spots list
import ast

# Extract coords and photos
blocks = re.split(r'\{\s*id:\s*', js_text)[1:]
for b in blocks:
    sid_m = re.match(r"['\"]([^'\"]+)['\"]", b)
    if sid_m:
        sid = sid_m.group(1)
        lat_m = re.search(r"lat:\s*([0-9.-]+)", b)
        lng_m = re.search(r"lng:\s*([0-9.-]+)", b)
        photos_m = re.search(r"photos:\s*\[(.*?)\]", b, re.DOTALL)
        
        lat = float(lat_m.group(1)) if lat_m else 30.2428
        lng = float(lng_m.group(1)) if lng_m else 120.1504
        photos = []
        if photos_m:
            raw_photos = photos_m.group(1)
            photos = [p.strip().strip("'\"") for p in raw_photos.split(',') if p.strip().strip("'\"")]
        
        if sid not in existing_spots_map:
            existing_spots_map[sid] = {'lat': lat, 'lng': lng, 'photos': photos}

DEFAULT_COORDS_BY_NUM = {
    96: (39.9163, 116.3971), # gugong forbidden city
    97: (39.8822, 116.4066), # tiantan
    98: (40.4319, 116.5704), # badaling great wall
    99: (30.2428, 120.1504), # westlake scenic
    100: (30.1318, 118.1691), # huangshan
    101: (36.2559, 117.1065), # taishan
    102: (34.4981, 110.0825), # huashan
    103: (25.2736, 110.2902), # guilin lijiang
    104: (29.3562, 110.4784), # zhangjiajie
    105: (36.7865, 99.0782),  # chaka
    106: (33.2600, 103.9186), # jiuzhaigou
    107: (48.7188, 87.0375),  # kanas
    108: (40.0360, 94.8020),  # dunhuang
    109: (29.6578, 91.1172),  # potala
    110: (31.3142, 120.6309), # suzhou gardens
    111: (30.7447, 120.4842), # wuzhen
    112: (25.7100, 100.2600), # dali erhai
    113: (26.8721, 100.2297), # lijiang
    114: (29.5630, 106.5516), # chongqing hongyadong
    115: (31.2402, 121.4905), # shanghai the bund
    116: (48.8606, 2.3376),   # louvre paris
    117: (51.5007, -0.1246),  # big ben london
    118: (40.7484, -73.9857), # empire state new york
    119: (35.6586, 139.7454), # tokyo tower
    120: (37.5512, 126.9882), # namsan seoul
    121: (1.2868, 103.8545),  # marina bay singapore
    122: (25.1972, 55.2744),  # burj khalifa dubai
    123: (29.9792, 31.1342),  # pyramids giza
    124: (-13.1631, -72.5450), # machu picchu
    125: (-22.9519, -43.2105), # christ rio
    126: (-20.1338, -67.4891), # salar de uyuni
    127: (41.8902, 12.4922),  # colosseum rome
    128: (27.1751, 78.0421),  # taj mahal
    129: (13.4125, 103.8670), # angkor wat
    130: (38.6431, 34.8289),  # cappadocia
    131: (-2.3333, 34.8333),  # serengeti
    132: (-16.5004, -151.7415), # bora bora
    133: (-65.0667, -64.0000)  # antarctica lemaire
}

wb = openpyxl.load_workbook(EXCEL_PATH)
sheet = wb.active

header_row_idx = 4
spots_list = []
seen_ids = set()

for r in range(header_row_idx + 1, sheet.max_row + 1):
    vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    if vals[0] is None:
        continue
    num = int(vals[0])
    raw_id = str(vals[1]).strip()
    name = str(vals[2]).strip()
    en_name = str(vals[3]).strip() if vals[3] else ""
    country = str(vals[4]).strip() if vals[4] else ""
    location = str(vals[5]).strip() if vals[5] else ""
    raw_category = str(vals[6]).strip() if vals[6] else ""
    desc = str(vals[7]).strip() if vals[7] else ""
    tags_str = str(vals[8]).strip() if vals[8] else ""
    bpm = vals[9] or 72
    scale = str(vals[10]).strip() if vals[10] else ""
    instruments = str(vals[11]).strip() if vals[11] else ""
    natural_sound = str(vals[12]).strip() if vals[12] else ""
    ai_prompt_zh = str(vals[13]).strip() if vals[13] else ""
    ai_prompt_en = str(vals[14]).strip() if vals[14] else ""

    spot_id = raw_id
    if spot_id in seen_ids:
        spot_id = f"{raw_id}_{num}"
    seen_ids.add(spot_id)

    cat_key = 'town'
    for k, v in CATEGORY_KEY_MAP.items():
        if k in raw_category:
            cat_key = v
            break

    tags = [t.strip() for t in re.split(r'[,、，|/]', tags_str) if t.strip()]

    # Coordinates
    lat, lng = DEFAULT_COORDS_BY_NUM.get(num, (30.2428, 120.1504))
    if raw_id in existing_spots_map:
        lat = existing_spots_map[raw_id]['lat']
        lng = existing_spots_map[raw_id]['lng']
    elif spot_id in existing_spots_map:
        lat = existing_spots_map[spot_id]['lat']
        lng = existing_spots_map[spot_id]['lng']

    # Photos
    photos = gpt2_spot_map.get(num, [])
    if not photos:
        if raw_id in existing_spots_map and existing_spots_map[raw_id]['photos']:
            photos = existing_spots_map[raw_id]['photos']
        elif spot_id in existing_spots_map and existing_spots_map[spot_id]['photos']:
            photos = existing_spots_map[spot_id]['photos']
        else:
            photos = ["https://images.unsplash.com/photo-1507525428034-b723cf961d3e?w=1200&auto=format&fit=crop&q=80"]

    spots_list.append({
        "id": spot_id,
        "rawId": raw_id,
        "name": name,
        "enName": en_name or name,
        "location": location,
        "country": country,
        "category": cat_key,
        "lat": lat,
        "lng": lng,
        "description": desc,
        "tags": tags,
        "photos": photos,
        "audioRecipe": {
            "style": "regional_acoustic",
            "bpm": int(bpm) if isinstance(bpm, (int, float)) else 72,
            "scale": scale,
            "instruments": instruments,
            "naturalSound": natural_sound,
            "aiPromptZh": ai_prompt_zh,
            "aiPromptEn": ai_prompt_en
        }
    })

# Extract 25 unclaimed ocean spots from existing spots.js
ocean_spots_js = []
ocean_matches = re.finditer(r"\{\s*id:\s*['\"]([a-z0-9_-]+)['\"].*?category:\s*['\"]unclaimed['\"].*?\}", js_text, re.DOTALL)
for m in ocean_matches:
    ocean_spots_js.append(m.group(0))

print(f"Extracted {len(ocean_spots_js)} ocean spot blocks.")

# Build spots.js content
out_lines = [
    "/**",
    " * GeoMelody Global Scenic Spots Dataset (133 Master Spots + 25 Ocean Spots = 158 Total)",
    " * Fully Synchronized with Excel Prompts, 133 Suno AI Music Tracks, and GPT2 Wallpapers",
    " */",
    "",
    "export const SCENIC_SPOTS = ["
]

for s in spots_list:
    out_lines.append("  {")
    out_lines.append(f"    id: {json.dumps(s['id'])},")
    if s['rawId'] != s['id']:
        out_lines.append(f"    rawId: {json.dumps(s['rawId'])},")
    out_lines.append(f"    name: {json.dumps(s['name'], ensure_ascii=False)},")
    out_lines.append(f"    enName: {json.dumps(s['enName'], ensure_ascii=False)},")
    out_lines.append(f"    location: {json.dumps(s['location'], ensure_ascii=False)},")
    out_lines.append(f"    country: {json.dumps(s['country'], ensure_ascii=False)},")
    out_lines.append(f"    category: {json.dumps(s['category'])},")
    out_lines.append(f"    lat: {s['lat']},")
    out_lines.append(f"    lng: {s['lng']},")
    out_lines.append(f"    description: {json.dumps(s['description'], ensure_ascii=False)},")
    out_lines.append(f"    tags: {json.dumps(s['tags'], ensure_ascii=False)},")
    out_lines.append(f"    photos: {json.dumps(s['photos'], ensure_ascii=False)},")
    out_lines.append("    audioRecipe: {")
    out_lines.append(f"      style: 'regional_acoustic',")
    out_lines.append(f"      bpm: {s['audioRecipe']['bpm']},")
    out_lines.append(f"      scale: {json.dumps(s['audioRecipe']['scale'], ensure_ascii=False)},")
    out_lines.append(f"      instruments: {json.dumps(s['audioRecipe']['instruments'], ensure_ascii=False)},")
    out_lines.append(f"      naturalSound: {json.dumps(s['audioRecipe']['naturalSound'], ensure_ascii=False)},")
    out_lines.append(f"      aiPromptZh: {json.dumps(s['audioRecipe']['aiPromptZh'], ensure_ascii=False)},")
    out_lines.append(f"      aiPromptEn: {json.dumps(s['audioRecipe']['aiPromptEn'], ensure_ascii=False)}")
    out_lines.append("    }")
    out_lines.append("  },")

# Append ocean spots
for o in ocean_spots_js:
    # Indent slightly
    out_lines.append("  " + o.strip() + ",")

out_lines.append("];")
out_lines.append("")

with open(os.path.join(DATA_DIR, 'spots.js'), 'w', encoding='utf-8') as f:
    f.write("\n".join(out_lines))

print(f"Generated clean src/data/spots.js with {len(spots_list)} + {len(ocean_spots_js)} spots.")
